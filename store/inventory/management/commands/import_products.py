"""
Comando Django para importar productos desde Excel/CSV v2

Columnas requeridas:
- Proveedor (obligatorio)
- Nombre Producto (obligatorio)
- Costo (obligatorio)
- Cantidad (obligatorio)
- % Minorista (obligatorio)
- % Mayorista (obligatorio)
- Marca (opcional)
- Descripcion (opcional)

Uso:
    python manage.py import_products archivo.csv
    python manage.py import_products archivo.xlsx
    python manage.py import_products archivo.xlsx --actualizar
"""

from django.core.management.base import BaseCommand
from inventory.models import Products, Category
from purchase.models import Supplier, Purchase, PurchaseProduct
from decimal import Decimal
from django.utils import timezone
import csv
import os
import random
import string


class Command(BaseCommand):
    help = 'Importa productos desde un archivo CSV o Excel v2 (con proveedor y marca)'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo CSV/Excel')
        parser.add_argument(
            '--actualizar',
            action='store_true',
            help='Actualizar productos existentes por nombre'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        actualizar = options.get('actualizar', False)

        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return

        # Determinar formato
        if archivo.endswith('.csv'):
            productos = self.leer_csv(archivo)
        elif archivo.endswith('.xlsx'):
            productos = self.leer_excel(archivo)
        else:
            self.stdout.write(self.style.ERROR('Solo se soportan archivos .csv y .xlsx'))
            return

        # Agrupar por proveedor para crear una Purchase por proveedor
        productos_por_proveedor = {}
        for producto_data in productos:
            proveedor_nombre = producto_data.get('proveedor', 'Sin Proveedor')
            if proveedor_nombre not in productos_por_proveedor:
                productos_por_proveedor[proveedor_nombre] = []
            productos_por_proveedor[proveedor_nombre].append(producto_data)

        # Procesar por proveedor
        creados = 0
        actualizados = 0
        errores = 0

        for proveedor_nombre, lista_productos in productos_por_proveedor.items():
            self.stdout.write(f"\n📦 Procesando proveedor: {proveedor_nombre}")

            # Obtener o crear proveedor
            proveedor, proveedor_creado = Supplier.objects.get_or_create(
                name=proveedor_nombre,
                defaults={'contact_info': ''}
            )
            if proveedor_creado:
                self.stdout.write(f"  ✨ Proveedor creado: {proveedor_nombre}")

            # Crear una Purchase para este proveedor
            purchase = Purchase.objects.create(
                supplier=proveedor,
                date_added=timezone.now(),
                pagado=False,
            )
            self.stdout.write(f"  📋 Compra #{purchase.id} creada para {proveedor_nombre}")

            # Procesar cada producto
            for row_num, producto_data in enumerate(lista_productos, start=2):
                try:
                    resultado, producto = self.crear_o_actualizar_producto(
                        producto_data, actualizar
                    )

                    if resultado == 'creado':
                        creados += 1
                        self.stdout.write(f"  ✅ {producto_data['nombre']} creado")
                    elif resultado == 'actualizado':
                        actualizados += 1
                        self.stdout.write(f"  🔄 {producto_data['nombre']} actualizado")
                    else:
                        self.stdout.write(f"  ⏭️  {producto_data['nombre']} ya existe (usar --actualizar)")
                        producto = Products.objects.get(name=producto_data['nombre'])

                    # Crear PurchaseProduct para vincular producto al proveedor
                    if producto and resultado in ('creado', 'actualizado'):
                        costo = Decimal(str(producto_data['costo'].replace(',', '.')))
                        cantidad = int(float(producto_data['cantidad'].replace(',', '.')))

                        PurchaseProduct.objects.create(
                            purchase=purchase,
                            product=producto,
                            cost=costo,
                            qty=cantidad,
                            total=costo * cantidad
                        )

                except Exception as e:
                    errores += 1
                    self.stdout.write(self.style.ERROR(
                        f"  ❌ {producto_data.get('nombre', '?')}: {str(e)}"
                    ))

            # Calcular total de la compra
            total_compra = sum(
                pp.total for pp in PurchaseProduct.objects.filter(purchase=purchase)
            )
            purchase.total = total_compra
            purchase.save()
            self.stdout.write(f"  💰 Total compra: AR$ {total_compra}")

        # Resumen final
        self.stdout.write(self.style.SUCCESS('\n=== RESUMEN ==='))
        self.stdout.write(self.style.SUCCESS(f'✅ Creados: {creados}'))
        if actualizar:
            self.stdout.write(self.style.SUCCESS(f'🔄 Actualizados: {actualizados}'))
        self.stdout.write(self.style.ERROR(f'❌ Errores: {errores}'))
        self.stdout.write(self.style.SUCCESS(
            f'📊 Total procesados: {creados + actualizados + errores}'
        ))
        self.stdout.write(self.style.SUCCESS(
            f'📦 Proveedores procesados: {len(productos_por_proveedor)}'
        ))

    def leer_csv(self, archivo):
        """Lee archivo CSV y retorna lista de productos"""
        productos = []
        with open(archivo, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                productos.append(self.normalizar_row(row))
        return productos

    def leer_excel(self, archivo):
        """Lee archivo Excel y retorna lista de productos"""
        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR(
                'Instala: pip install openpyxl --break-system-packages'
            ))
            return []

        productos = []
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, row))
            productos.append(self.normalizar_row(data))

        return productos

    def normalizar_row(self, data):
        """Normaliza los datos de una fila (acepta variantes de nombres de columna)"""
        return {
            'proveedor': str(data.get('Proveedor', data.get('proveedor', 'Sin Proveedor')) or 'Sin Proveedor').strip(),
            'nombre': str(data.get('Nombre Producto', data.get('nombre', '')) or '').strip(),
            'costo': str(data.get('Costo', data.get('costo', '0')) or '0').strip(),
            'cantidad': str(data.get('Cantidad', data.get('cantidad', '0')) or '0').strip(),
            'porcentaje_minorista': str(data.get('% Minorista', data.get('% minorista', '0')) or '0').strip(),
            'porcentaje_mayorista': str(data.get('% Mayorista', data.get('% mayorista', '0')) or '0').strip(),
            'marca': str(data.get('Marca', data.get('marca', '')) or '').strip(),
            'descripcion': str(data.get('Descripcion', data.get('Descripción', '')) or '').strip(),
        }

    def generar_codigo_unico(self):
        """Genera un código único para el producto"""
        while True:
            codigo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not Products.objects.filter(code=codigo).exists():
                return codigo

    def crear_o_actualizar_producto(self, data, actualizar=False):
        """Crea o actualiza un producto"""

        if not data['nombre']:
            raise ValueError('Nombre de producto es obligatorio')

        # Calcular precios
        costo = round(Decimal(data['costo'].replace(',', '.')), 2)
        porc_minorista = Decimal(data['porcentaje_minorista'].replace(',', '.'))
        porc_mayorista = Decimal(data['porcentaje_mayorista'].replace(',', '.'))

        precio_minorista = round(costo * (1 + porc_minorista / 100), 2)
        precio_mayorista = round(costo * (1 + porc_mayorista / 100), 2)

        cantidad = int(float(data['cantidad'].replace(',', '.')))

        # Obtener o crear categoría usando el proveedor como categoría
        categoria, _ = Category.objects.get_or_create(
            name=data['proveedor'],
            defaults={'status': 1}
        )

        # Verificar si existe
        producto_existente = Products.objects.filter(name=data['nombre']).first()

        if producto_existente:
            if actualizar:
                producto_existente.cost = costo
                producto_existente.precio_minorista = precio_minorista
                producto_existente.precio_mayorista = precio_mayorista
                producto_existente.margen_minorista = porc_minorista
                producto_existente.margen_mayorista = porc_mayorista
                producto_existente.quantity = cantidad
                if data.get('marca'):
                    producto_existente.marca = data['marca']
                if data.get('descripcion'):
                    producto_existente.description = data['descripcion']
                producto_existente.save()
                return 'actualizado', producto_existente
            else:
                return 'existe', None
        else:
            # Crear nuevo
            producto = Products.objects.create(
                code=self.generar_codigo_unico(),
                category=categoria,
                name=data['nombre'],
                cost=costo,
                precio_minorista=precio_minorista,
                precio_mayorista=precio_mayorista,
                margen_minorista=porc_minorista,
                margen_mayorista=porc_mayorista,
                quantity=cantidad,
                marca=data.get('marca', ''),
                description=data.get('descripcion', ''),
                status=1
            )
            return 'creado', producto
